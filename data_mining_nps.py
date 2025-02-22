# -*- coding: utf-8 -*-
"""data-mining_NPS

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1lPZ6bXo8lw_lHObEPSJW61MuCORLmNNR
"""

from google.colab import drive
drive.mount('/content/drive')

from collections import Counter
import openpyxl

# Load the file
workbook = openpyxl.load_workbook('/content/drive/My Drive/Caso.xlsx')

# Select the desired spreadsheet
sheet = workbook['texto']

# Initializes a word counter
word_counter = Counter() # Define um número mínimo de caracteres para uma palavra
min_word_length = 8

# Function to check if a word contains only alphabetic characters
def is_alpha(word):
    return word.isalpha()

# Function to check if two words are singular and plural variations of each other
def is_singular_plural(word1, word2):
    if len(word1) < 2 or len(word2) < 2:
        return False
    if word1[-1] == 's' and word1[:-1] == word2:
        return True
    elif word2[-1] == 's' and word2[:-1] == word1:
        return True
    return False

# Function to check if two words are variations of each other's masculine and feminine forms
def is_male_female(word1, word2):
    if len(word1) < 2 or len(word2) < 2:
        return False
    if word1[-1] == 'o' and word1[:-1] == word2:
        return True
    elif word2[-1] == 'a' and word2[:-1] == word1:
        return True
    return False

# Iterates over the cells in the column with the comments
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
    comment = row[0]
    if isinstance(comment, str):  # Checks if the comment is a string
        words = comment.lower().split()  # Convert to lowercase and split the comment into words
        for word in words:
            if len(word) >= min_word_length and is_alpha(word):  # Checks minimum number of characters alphabetical format
                duplicated = False
                for key in word_counter.keys():
                    if is_singular_plural(word, key) or is_male_female(word, key):
                        word_counter[key] += 1
                        duplicated = True
                        break
                if not duplicated:
                    word_counter[word] += 1

# Print the 15 most common words
print(word_counter.most_common(15))

workbook.close()